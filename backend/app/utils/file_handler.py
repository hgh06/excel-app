import os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from utils.config import UPLOAD_FOLDER, ALLOWED_EXTENSIONS
from utils.helpers import unmerge_cells, delete_rows, delete_columns, delete_empty_rows, set_borders_and_alignment, adjust_row_heights

def allowed_file(filename):
    """Check if the file has a valid extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_file(file):
    """Save the uploaded file to the server."""
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(file_path)
        return file_path
    return None

def load_workbook_and_sheet(input_file, sheet_name):
    wb = load_workbook(input_file)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} does not exist in the workbook")
    
    return wb, wb[sheet_name]

def modify_excel(file_path):
    """Modify the uploaded Excel file."""
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Select the active sheet (or specify sheet by name)
        ws = wb['Original']
        unmerge_cells(ws)
        delete_rows(ws)
        delete_columns(ws)
        delete_empty_rows(ws)
        set_borders_and_alignment(ws)
        adjust_row_heights(ws)
        
        # Save the modified workbook with a new name
        modified_file_path = os.path.join(UPLOAD_FOLDER, 'modified_' + os.path.basename(file_path))
        wb.save(modified_file_path)
        
        return modified_file_path
    except Exception as e:
        raise e