from openpyxl.styles import Border, Side, Alignment
from .excel_config import DEFAULT_TITLE, DEFAULT_COLUMNS_TO_KEEP

def unmerge_cells(ws):
    # unmerge cells
    for merged_cell in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_cell))

def delete_rows(ws, delete_until_value=DEFAULT_TITLE):
    # delete fixed template part at the top
    ws.delete_rows(1,33)

    # delete rows starting with fixed title
    for row in range(ws.max_row, 0, -1):
        cell_value = ws.cell(row=row, column=1).value
        if isinstance(cell_value, str) and delete_until_value == cell_value:
            ws.delete_rows(row, ws.max_row)
            break

def delete_columns(ws, columns_to_keep=DEFAULT_COLUMNS_TO_KEEP):
    # delete not necessaty columns
    first_row = [cell.value for cell in ws[1]]
    for col in range(ws.max_column, 0, -1):
        column_title = first_row[col - 1]
        if column_title not in columns_to_keep:
            ws.delete_cols(col)

def delete_empty_rows(ws):
    # delete empty rows
    for row in range(ws.max_row, 0, -1):
        is_empty = True
        for cell in ws[row]:
            if cell.value is not None:
                is_empty = False
                break
        if is_empty:
            ws.delete_rows(row)

def set_borders_and_alignment(ws):
    # Define border style
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
        
    # Set borders for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.border = border

def adjust_row_heights(ws):
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                try:
                    cell_value = str(cell.value)
                    lines = cell_value.split('\n')
                    max_height = max(max_height, len(lines))
                except:
                    pass
        ws.row_dimensions[row[0].row].height = max_height * 15